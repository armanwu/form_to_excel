from tkinter import *
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

# Fungsi untuk menyimpan data ke file Excel
def simpan_ke_excel():
    # Membuka file Excel yang sudah ada atau membuat baru jika belum ada
    try:
        wb = load_workbook("data_form.xlsx")
        sheet = wb.active
    except FileNotFoundError:
        wb = Workbook()
        sheet = wb.active
        sheet.append(["Nama", "Email", "Telepon"])

    # Menyimpan data dari form ke dalam file Excel
    data = []
    data.append(nama_entry.get())
    data.append(email_entry.get())
    data.append(telepon_entry.get())

    # Menambahkan data ke baris baru di file Excel
    row = sheet.max_row + 1
    sheet.append(data)

    # Menyimpan file Excel
    wb.save("data_form.xlsx")

    # Membersihkan form setelah data tersimpan
    nama_entry.delete(0, END)
    email_entry.delete(0, END)
    telepon_entry.delete(0, END)

    # Memberi tahu pengguna bahwa data telah tersimpan
    info_label.config(text="Data tersimpan di data_form.xlsx", fg="green")

# Membuat GUI form menggunakan tkinter
root = Tk()
root.title("Form Data")
root.geometry("300x150")

# Label dan Entry untuk Nama
nama_label = Label(root, text="Nama:")
nama_label.pack()
nama_entry = Entry(root)
nama_entry.pack()

# Label dan Entry untuk Email
email_label = Label(root, text="Email:")
email_label.pack()
email_entry = Entry(root)
email_entry.pack()

# Label dan Entry untuk Telepon
telepon_label = Label(root, text="Telepon:")
telepon_label.pack()
telepon_entry = Entry(root)
telepon_entry.pack()

# Tombol Simpan
simpan_button = Button(root, text="Simpan", command=simpan_ke_excel)
simpan_button.pack()

# Label informasi
info_label = Label(root, text="")
info_label.pack()

root.mainloop()